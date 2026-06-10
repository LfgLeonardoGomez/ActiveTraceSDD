"""Servicio de gestión de guardias (C-13).

Reglas de negocio:
- TUTOR puede registrar guardias para sus propias asignaciones.
- COORDINADOR/ADMIN ven todas las guardias del tenant.
- Export CSV/XLSX reutiliza patrón de EquiposService (C-08).
- Audit log en todas las operaciones de escritura.

No accede directamente a DB: delega todo a repositories.
"""

import csv
import io
from datetime import date
from typing import Any
from uuid import UUID

from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit_log import AuditLog
from app.models.guardia import Guardia
from app.repositories.audit_log_repository import AuditLogRepository
from app.repositories.guardias import GuardiaRepository
from app.schemas.guardias import (
    ExportarGuardiasParams,
    GuardiaCreate,
    GuardiaFilterParams,
    GuardiaRead,
    GuardiaUpdate,
    PaginatedGuardiaResponse,
)

# Opcional: openpyxl para XLSX
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False


class GuardiaService:
    """Servicio de gestión de guardias del tenant."""

    def __init__(self, db_session: AsyncSession, tenant_id: UUID) -> None:
        self.db_session = db_session
        self.tenant_id = tenant_id
        self._repo = GuardiaRepository(db_session, tenant_id)
        self._repo_audit = AuditLogRepository(db_session, tenant_id)

    async def _insert_audit_log(
        self,
        actor_id: UUID,
        accion: str,
        detalle: dict[str, Any],
        filas_afectadas: int = 0,
        materia_id: UUID | None = None,
    ) -> None:
        """Inserta registro de auditoría en la transacción activa."""
        entry = AuditLog(
            tenant_id=self.tenant_id,
            actor_id=actor_id,
            accion=accion,
            detalle=detalle,
            filas_afectadas=filas_afectadas,
            materia_id=materia_id,
        )
        await self._repo_audit.insert(entry)

    def _map_row_to_guardia_read(self, row: tuple[Any, ...]) -> GuardiaRead:
        """Mapea una tupla de resultado JOIN a GuardiaRead."""
        guardia = row[0]
        return GuardiaRead.model_validate(
            {
                "id": guardia.id,
                "tenant_id": guardia.tenant_id,
                "tutor_id": guardia.tutor_id,
                "materia_id": guardia.materia_id,
                "carrera_id": guardia.carrera_id,
                "cohorte_id": guardia.cohorte_id,
                "fecha": guardia.fecha,
                "horario": guardia.horario,
                "descripcion": guardia.descripcion,
                "estado": guardia.estado,
                "comentarios": guardia.comentarios,
                "created_at": str(guardia.created_at) if guardia.created_at else None,
                "updated_at": str(guardia.updated_at) if guardia.updated_at else None,
            }
        )

    async def registrar_guardia(
        self,
        data: GuardiaCreate,
        actor_id: UUID,
    ) -> Guardia:
        """Registra una nueva guardia.

        Args:
            data: datos de la guardia.
            actor_id: UUID del usuario que registra la guardia.

        Returns:
            Guardia creada.
        """
        guardia = await self._repo.create(
            tutor_id=actor_id,
            materia_id=data.materia_id,
            carrera_id=data.carrera_id,
            cohorte_id=data.cohorte_id,
            fecha=data.fecha,
            horario=data.horario,
            descripcion=data.descripcion,
            estado="Pendiente",
        )

        await self._insert_audit_log(
            actor_id=actor_id,
            accion="GUARDIA_REGISTRAR",
            detalle={
                "guardia_id": str(guardia.id),
                "materia_id": str(data.materia_id),
                "fecha": str(data.fecha),
            },
            filas_afectadas=1,
            materia_id=data.materia_id,
        )

        return guardia

    async def listar_guardias(
        self,
        filters: GuardiaFilterParams,
        limit: int,
        offset: int,
        current_user_id: UUID,
        current_user_roles: list[str],
    ) -> PaginatedGuardiaResponse:
        """Lista guardias paginadas con filtros.

        Scope:
        - COORDINADOR / ADMIN: ven todas las guardias del tenant.
        - TUTOR / PROFESOR: solo las propias.
        """
        # Scope por rol: TUTOR/PROFESOR solo ven las propias
        effective_tutor_id: UUID | None = None
        if not any(r in ("ADMIN", "COORDINADOR") for r in current_user_roles):
            effective_tutor_id = current_user_id

        tutor_filter = filters.tutor_id or effective_tutor_id

        items, total = await self._repo.list_guardias(
            materia_id=filters.materia_id,
            tutor_id=tutor_filter,
            estado=filters.estado,
            fecha_desde=filters.fecha_desde,
            fecha_hasta=filters.fecha_hasta,
            limit=limit,
            offset=offset,
        )

        return PaginatedGuardiaResponse(
            items=[self._map_row_to_guardia_read(row) for row in items],
            total=total,
            limit=limit,
            offset=offset,
        )

    async def obtener_guardia(self, guardia_id: UUID) -> Guardia:
        """Obtiene una guardia por ID.

        Raises:
            HTTPException 404: si no existe.
        """
        guardia = await self._repo.get_by_id(guardia_id)
        if guardia is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Guardia no encontrada",
            )
        return guardia

    async def actualizar_guardia(
        self,
        guardia_id: UUID,
        data: GuardiaUpdate,
        actor_id: UUID,
    ) -> Guardia:
        """Actualiza una guardia.

        Raises:
            HTTPException 404: si no existe.
        """
        update_data = data.model_dump(exclude_unset=True)
        if not update_data:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No se proporcionaron campos para actualizar",
            )

        guardia = await self._repo.update(guardia_id, update_data)
        if guardia is None:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Guardia no encontrada",
            )

        await self._insert_audit_log(
            actor_id=actor_id,
            accion="GUARDIA_REGISTRAR",
            detalle={
                "guardia_id": str(guardia_id),
                "campos": list(update_data.keys()),
            },
            filas_afectadas=1,
            materia_id=guardia.materia_id,
        )

        return guardia

    async def exportar_guardias(
        self,
        params: ExportarGuardiasParams,
        current_user_id: UUID,
        current_user_roles: list[str],
    ) -> StreamingResponse:
        """Exporta guardias a CSV o XLSX.

        Scope:
        - COORDINADOR / ADMIN: exportan todas las guardias del tenant.
        - TUTOR / PROFESOR: solo las propias.
        """
        effective_tutor_id: UUID | None = None
        if not any(r in ("ADMIN", "COORDINADOR") for r in current_user_roles):
            effective_tutor_id = current_user_id

        rows = await self._repo.get_guardias_for_export(
            materia_id=params.materia_id,
            tutor_id=effective_tutor_id,
            estado=params.estado,
            fecha_desde=params.fecha_desde,
            fecha_hasta=params.fecha_hasta,
        )

        headers = [
            "Tutor",
            "Materia",
            "Carrera",
            "Cohorte",
            "Fecha",
            "Horario",
            "Estado",
            "Descripcion",
        ]

        data_rows = []
        for row in rows:
            guardia = row[0]
            data_rows.append(
                [
                    f"{row[1]} {row[2]}",
                    row[3],
                    row[4],
                    row[5],
                    str(guardia.fecha),
                    guardia.horario or "",
                    guardia.estado,
                    guardia.descripcion,
                ]
            )

        if params.formato == "xlsx":
            if not _HAS_OPENPYXL:
                raise HTTPException(
                    status_code=status.HTTP_501_NOT_IMPLEMENTED,
                    detail="openpyxl no disponible para exportar XLSX",
                )

            wb = Workbook()
            ws = wb.active
            ws.title = "Guardias"
            ws.append(headers)
            for r in data_rows:
                ws.append(r)

            # Ajustar anchos de columna
            for i, h in enumerate(headers, 1):
                ws.column_dimensions[get_column_letter(i)].width = max(len(h), 15)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            return StreamingResponse(
                buffer,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": "attachment; filename=guardias.xlsx"
                },
            )

        # CSV por defecto
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(data_rows)

        return StreamingResponse(
            io.BytesIO(buffer.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=guardias.csv"},
        )
