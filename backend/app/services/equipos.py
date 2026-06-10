"""Servicio de gestión de equipos docentes (C-08).

Reglas de negocio:
- Asignación masiva: valida todos los usuario_ids antes de crear; atómica.
- Clonación: solo copia asignaciones vigentes; originales intactos.
- Vigencia: batch update solo de asignaciones vigentes; 422 si no hay.
- Export: excluye PII por defecto; requiere permiso equipos:ver-pii.
- Audit log en todas las operaciones de escritura.

No accede directamente a DB: delega todo a repositories.
"""

import csv
import io
from datetime import date
from typing import Any
from uuid import UUID

from fastapi import HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit_log import AuditLog
from app.repositories.asignaciones import AsignacionRepository
from app.repositories.audit_log_repository import AuditLogRepository
from app.repositories.equipos import EquipoRepository
from app.repositories.usuarios import UsuarioRepository
from app.schemas.equipos import (
    AsignacionMasivaRequest,
    AsignacionMasivaResponse,
    ActualizarVigenciaRequest,
    ActualizarVigenciaResponse,
    ClonarEquipoRequest,
    ClonarEquipoResponse,
    EquipoFilterParams,
    EquipoRead,
    PaginatedEquipoResponse,
)

# Opcional: openpyxl para XLSX
try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:  # pragma: no cover
    _HAS_OPENPYXL = False


class EquipoService:
    """Servicio de gestión de equipos docentes del tenant."""

    def __init__(self, db_session: AsyncSession, tenant_id: UUID) -> None:
        self.db_session = db_session
        self.tenant_id = tenant_id
        self._repo_equipo = EquipoRepository(db_session, tenant_id)
        self._repo_asig = AsignacionRepository(db_session, tenant_id)
        self._repo_user = UsuarioRepository(db_session, tenant_id)
        self._repo_audit = AuditLogRepository(db_session, tenant_id)

    def _map_row_to_equipo_read(self, row: tuple[Any, ...]) -> EquipoRead:
        """Mapea una tupla de resultado JOIN a EquipoRead."""
        asignacion = row[0]
        return EquipoRead.model_validate(
            {
                "id": asignacion.id,
                "tenant_id": asignacion.tenant_id,
                "usuario_id": asignacion.usuario_id,
                "rol": asignacion.rol,
                "desde": asignacion.desde,
                "hasta": asignacion.hasta,
                "materia_id": asignacion.materia_id,
                "carrera_id": asignacion.carrera_id,
                "cohorte_id": asignacion.cohorte_id,
                "comisiones": asignacion.comisiones,
                "responsable_id": asignacion.responsable_id,
                "estado_vigencia": asignacion.estado_vigencia,
                "materia_nombre": row[1] if len(row) > 1 else None,
                "carrera_nombre": row[2] if len(row) > 2 else None,
                "cohorte_nombre": row[3] if len(row) > 3 else None,
                "usuario_nombre": row[4] if len(row) > 4 else None,
                "usuario_apellidos": row[5] if len(row) > 5 else None,
            }
        )

    async def _insert_audit_log(
        self,
        actor_id: UUID,
        accion: str,
        detalle: dict[str, Any],
        filas_afectadas: int,
        materia_id: UUID | None = None,
        impersonado_id: UUID | None = None,
    ) -> None:
        """Inserta registro de auditoría en la transacción activa."""
        entry = AuditLog(
            tenant_id=self.tenant_id,
            actor_id=actor_id,
            accion=accion,
            detalle=detalle,
            filas_afectadas=filas_afectadas,
            materia_id=materia_id,
            impersonado_id=impersonado_id,
        )
        await self._repo_audit.insert(entry)

    async def mis_equipos(
        self,
        current_user,
        filters: EquipoFilterParams,
        limit: int,
        offset: int,
    ) -> PaginatedEquipoResponse:
        """Lista asignaciones del usuario autenticado con contexto académico."""
        items, total = await self._repo_equipo.list_by_usuario(
            usuario_id=current_user.id,
            estado_vigencia=filters.estado_vigencia,
            materia_id=filters.materia_id,
            carrera_id=filters.carrera_id,
            cohorte_id=filters.cohorte_id,
            limit=limit,
            offset=offset,
        )
        return PaginatedEquipoResponse(
            items=[self._map_row_to_equipo_read(row) for row in items],
            total=total,
            limit=limit,
            offset=offset,
        )

    async def list_equipo(
        self,
        materia_id: UUID,
        carrera_id: UUID,
        cohorte_id: UUID,
        estado_vigencia: str | None,
        limit: int,
        offset: int,
    ) -> PaginatedEquipoResponse:
        """Lista asignaciones de un equipo por contexto académico."""
        items, total = await self._repo_equipo.list_by_equipo(
            materia_id=materia_id,
            carrera_id=carrera_id,
            cohorte_id=cohorte_id,
            estado_vigencia=estado_vigencia,
            limit=limit,
            offset=offset,
        )
        return PaginatedEquipoResponse(
            items=[self._map_row_to_equipo_read(row) for row in items],
            total=total,
            limit=limit,
            offset=offset,
        )

    async def asignacion_masiva(
        self,
        data: AsignacionMasivaRequest,
        current_user,
    ) -> AsignacionMasivaResponse:
        """Asignación masiva de docentes a un equipo.

        Valida que todos los usuario_ids existan antes de crear.
        Atómica: si falla, ninguna fila se crea.
        """
        # Validar que todos los usuarios existan
        for uid in data.usuario_ids:
            usuario = await self._repo_user.get_by_id(uid)
            if usuario is None:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail=f"usuario_id {uid} no encontrado",
                )

        items = []
        for uid in data.usuario_ids:
            items.append(
                {
                    "usuario_id": uid,
                    "rol": data.rol,
                    "desde": data.desde,
                    "hasta": data.hasta,
                    "materia_id": data.materia_id,
                    "carrera_id": data.carrera_id,
                    "cohorte_id": data.cohorte_id,
                    "comisiones": [],
                    "responsable_id": None,
                }
            )

        created = await self._repo_equipo.bulk_create_assignments(items)
        await self.db_session.commit()

        created_ids = [c.id for c in created]
        detalle = {
            "materia_id": str(data.materia_id),
            "carrera_id": str(data.carrera_id),
            "cohorte_id": str(data.cohorte_id),
            "rol": data.rol,
        }
        await self._insert_audit_log(
            actor_id=current_user.real_actor_id,
            accion="ASIGNACION_CREAR",
            detalle=detalle,
            filas_afectadas=len(created),
            materia_id=data.materia_id,
            impersonado_id=current_user.impersonated_id,
        )
        await self.db_session.commit()

        return AsignacionMasivaResponse(
            count=len(created),
            created_ids=created_ids,
        )

    async def clonar_equipo(
        self,
        data: ClonarEquipoRequest,
        current_user,
    ) -> ClonarEquipoResponse:
        """Clona asignaciones vigentes de un equipo origen a destino.

        Si preview=True, solo cuenta sin crear.
        """
        if data.preview:
            # Contar vigentes sin crear
            today = date.today()
            count_query = await self._repo_asig.list_paginated(
                limit=1000,
                offset=0,
                materia_id=data.materia_id,
                carrera_id=data.carrera_id,
                cohorte_id=data.cohorte_id_origen,
                incluir_vencidas=False,
            )
            preview_count = count_query[1]
            return ClonarEquipoResponse(preview_count=preview_count)

        created = await self._repo_equipo.clone_vigentes(
            materia_id=data.materia_id,
            carrera_id=data.carrera_id,
            cohorte_id_origen=data.cohorte_id_origen,
            cohorte_id_destino=data.cohorte_id_destino,
            desde=data.desde,
            hasta=data.hasta,
        )

        if not created:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No hay asignaciones vigentes para clonar",
            )

        await self.db_session.commit()

        created_ids = [c.id for c in created]
        detalle = {
            "materia_id": str(data.materia_id),
            "carrera_id": str(data.carrera_id),
            "cohorte_id_origen": str(data.cohorte_id_origen),
            "cohorte_id_destino": str(data.cohorte_id_destino),
        }
        await self._insert_audit_log(
            actor_id=current_user.real_actor_id,
            accion="ASIGNACION_CLONAR",
            detalle=detalle,
            filas_afectadas=len(created),
            materia_id=data.materia_id,
            impersonado_id=current_user.impersonated_id,
        )
        await self.db_session.commit()

        return ClonarEquipoResponse(
            preview_count=len(created),
            created_count=len(created),
            created_ids=created_ids,
        )

    async def actualizar_vigencia(
        self,
        materia_id: UUID,
        carrera_id: UUID,
        cohorte_id: UUID,
        data: ActualizarVigenciaRequest,
        current_user,
    ) -> ActualizarVigenciaResponse:
        """Actualiza vigencia de todas las asignaciones vigentes del equipo."""
        count = await self._repo_equipo.update_vigencia_by_equipo(
            materia_id=materia_id,
            carrera_id=carrera_id,
            cohorte_id=cohorte_id,
            desde=data.desde,
            hasta=data.hasta,
        )

        if count == 0:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="No hay asignaciones vigentes en el equipo",
            )

        await self.db_session.commit()

        detalle = {
            "materia_id": str(materia_id),
            "carrera_id": str(carrera_id),
            "cohorte_id": str(cohorte_id),
            "desde": str(data.desde),
            "hasta": str(data.hasta) if data.hasta else None,
        }
        await self._insert_audit_log(
            actor_id=current_user.real_actor_id,
            accion="ASIGNACION_MODIFICAR",
            detalle=detalle,
            filas_afectadas=count,
            materia_id=materia_id,
            impersonado_id=current_user.impersonated_id,
        )
        await self.db_session.commit()

        return ActualizarVigenciaResponse(count=count)

    async def exportar_equipo(
        self,
        materia_id: UUID,
        carrera_id: UUID,
        cohorte_id: UUID,
        format: str,
        include_pii: bool,
        current_user,
    ) -> StreamingResponse:
        """Exporta equipo a CSV o XLSX.

        Si include_pii=True, verifica permiso equipos:ver-pii.
        """
        if include_pii:
            # Verificar permiso equipos:ver-pii
            from app.services.permission_service import PermissionService

            perm_service = PermissionService(self.db_session, self.tenant_id)
            effective = await perm_service.resolve_effective_permissions(
                current_user.roles
            )
            perm_map = {code: propio for code, propio in effective}
            if "equipos:ver-pii" not in perm_map:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail="Permiso denegado para ver PII",
                )

        rows = await self._repo_equipo.get_equipo_for_export(
            materia_id=materia_id,
            carrera_id=carrera_id,
            cohorte_id=cohorte_id,
            include_pii=include_pii,
        )

        # Preparar datos planos
        data_rows = []
        for row in rows:
            asig = row[0]
            flat = {
                "nombre": row[4] if len(row) > 4 else None,
                "apellidos": row[5] if len(row) > 5 else None,
                "rol": asig.rol,
                "materia": row[1] if len(row) > 1 else None,
                "carrera": row[2] if len(row) > 2 else None,
                "cohorte": row[3] if len(row) > 3 else None,
                "comisiones": ", ".join(asig.comisiones) if asig.comisiones else "",
                "desde": str(asig.desde),
                "hasta": str(asig.hasta) if asig.hasta else "",
                "estado_vigencia": asig.estado_vigencia,
            }
            if include_pii and len(row) > 8:
                flat["email"] = row[6]
                flat["dni"] = row[7]
                flat["cbu"] = row[8]
            data_rows.append(flat)

        filename = f"equipo_{materia_id}_{carrera_id}_{cohorte_id}"

        if format == "xlsx":
            if not _HAS_OPENPYXL:
                raise HTTPException(
                    status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail="openpyxl no está disponible",
                )
            output = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Equipo"
            headers = list(data_rows[0].keys()) if data_rows else [
                "nombre", "apellidos", "rol", "materia", "carrera",
                "cohorte", "comisiones", "desde", "hasta", "estado_vigencia",
            ]
            if include_pii:
                headers += ["email", "dni", "cbu"]
            # Escribir headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            # Escribir datos
            for row_idx, data_row in enumerate(data_rows, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col, value=data_row.get(header, ""))
            # Auto-width simple
            for col, header in enumerate(headers, 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = max(len(header) + 2, 12)
            wb.save(output)
            output.seek(0)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ext = "xlsx"
        else:
            # CSV default
            output = io.StringIO()
            writer = csv.DictWriter(
                output,
                fieldnames=[
                    "nombre", "apellidos", "rol", "materia", "carrera",
                    "cohorte", "comisiones", "desde", "hasta", "estado_vigencia",
                ] + (["email", "dni", "cbu"] if include_pii else []),
            )
            writer.writeheader()
            writer.writerows(data_rows)
            output.seek(0)
            media_type = "text/csv; charset=utf-8"
            ext = "csv"

        # Audit log
        detalle = {
            "materia_id": str(materia_id),
            "carrera_id": str(carrera_id),
            "cohorte_id": str(cohorte_id),
            "format": format,
            "include_pii": include_pii,
        }
        await self._insert_audit_log(
            actor_id=current_user.real_actor_id,
            accion="EQUIPO_EXPORTAR",
            detalle=detalle,
            filas_afectadas=len(data_rows),
            materia_id=materia_id,
            impersonado_id=current_user.impersonated_id,
        )
        await self.db_session.commit()

        return StreamingResponse(
            io.BytesIO(output.getvalue().encode("utf-8") if isinstance(output, io.StringIO) else output.getvalue()),
            media_type=media_type,
            headers={
                "Content-Disposition": f'attachment; filename="{filename}.{ext}"'
            },
        )
